from datetime import date, datetime, timedelta
import csv
import io
import os

from django.conf import settings as django_settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from django.core.files.storage import default_storage
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.utils.text import slugify
from django.views.decorators.csrf import csrf_exempt

import openpyxl
# from asgiref.sync import async_to_sync

from attendance.models import Student, Attendance, Course, Camera, AttendanceSchedule
from attendance.take_attendance import take_attendance


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('dashboard-home')
        else:
            return render(request, 'login.html', {'error': 'Invalid credentials'})
    return render(request, 'login.html')


@login_required(login_url='login')
def dashboard_home(request):
    selected_class_id = request.GET.get('class')
    today = date.today()
    week_days = [(today - timedelta(days=i)) for i in range(6, -1, -1)]
    week_days_labels = [d.strftime('%a') for d in week_days]

    # Filter students by class if selected
    students = Student.objects.all()
    if selected_class_id:
        students = students.filter(student_class_id=selected_class_id)

    total_students = students.count()
    today_attendance = Attendance.objects.filter(student__in=students, date=today, status='Present').count()
    absent_today = total_students - today_attendance

    # Attendance trend for the past 7 days
    attendance_trend = [
        Attendance.objects.filter(student__in=students, date=day, status='Present').count()
        for day in week_days
    ]

    # Pie chart data
    pie_data = [today_attendance, absent_today]

    # Top 3 students with best attendance (most presents)
    top_students = students.annotate(presents=Count('attendance', filter=Q(attendance__status='Present'))).order_by('-presents')[:3]
    # Top 3 frequent absentees
    frequent_absentees = students.annotate(absents=Count('attendance', filter=Q(attendance__status='Absent'))).order_by('-absents')[:3]

    classes = Course.objects.all()

    context = {
        'total_students': total_students,
        'today_attendance': today_attendance,
        'absent_today': absent_today,
        'attendance_trend': attendance_trend,
        'week_days_labels': week_days_labels,
        'pie_data': pie_data,
        'top_students': top_students,
        'frequent_absentees': frequent_absentees,
        'classes': classes,
        'selected_class_id': selected_class_id,
    }
    return render(request, 'contents/dashboard.html', context)


def logout_view(request):
    logout(request)
    return redirect('login')

@login_required(login_url='login')
def attendance(request):
    # Get filter params
    class_id = request.GET.get('class')
    date_str = request.GET.get('date')
    action = request.GET.get('action', '')
    filetype = request.GET.get('filetype', 'csv')
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except Exception:
            selected_date = date.today()
    else:
        selected_date = date.today()

    students = Student.objects.all().select_related('student_class')
    if class_id:
        students = students.filter(student_class_id=class_id)

    cameras = Camera.objects.all()
    message = None
    a_status = None

    # Handle download action
    if action == 'download':
        classroom_name = "all_classes"
        if class_id:
            classroom = Course.objects.filter(id=class_id).first()
            if classroom:
                classroom_name = classroom.name.replace(" ", "_")
        attendance_records = Attendance.objects.filter(student__in=students, date=selected_date)
        attendance_map = {a.student.id: a.status for a in attendance_records}
        filename = f"{classroom_name}_attendance_{selected_date}"
        headers = ['Student ID', 'Name', 'Roll No', 'Status']
        data_rows = [
            [student.id, student.name, student.roll_number, attendance_map.get(student.id, '')]
            for student in students
        ]
        if filetype == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(headers)
            for row in data_rows:
                ws.append(row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response
        else:  # default to CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            writer = csv.writer(response)
            writer.writerow(headers)
            for row in data_rows:
                writer.writerow(row)
            return response

    #POST: update attendance
    if request.method == 'POST':
        for student in students:
            status = request.POST.get(f'status_{student.id}')
            if status in ['Present', 'Absent']:
                Attendance.objects.update_or_create(
                    student=student,
                    date=selected_date,
                    defaults={'status': status}
                )

    # Handle manual take_attendance POST
    if request.method == 'POST' and request.POST.get('action') == 'take_attendance':
        course_id = request.POST.get('course')
        camera_id = request.POST.get('camera')
        date_val = request.POST.get('date') or selected_date
        a_status, result = take_attendance(camera_id, course_id, for_date=date_val)
        # result = async_to_sync(take_attendance)(camera_id, course_id, for_date=date_val)

        return redirect(f"{request.path}?class={course_id}&date={date_val}&status={a_status}&message={str(result)}")

    # Get attendance records for the selected date
    attendance_records = Attendance.objects.filter(student__in=students, date=selected_date)
    attendance_map = {a.student.id: a.status for a in attendance_records}
    for student in students:
        student.attendance_status = attendance_map.get(student.id, "")

    # For class dropdown
    classes = Course.objects.all()
    a_status= request.GET.get('status') or a_status
    message = request.GET.get('message') or message
    context = {
        'default_date': selected_date.isoformat(),
        'students': students,
        'attendance_map': attendance_map,
        'classes': classes,
        'selected_class_id': class_id or '',
        'cameras': cameras,
        "status": a_status,
        'message': message,
    }
    return render(request, 'contents/attendance.html', context)

@login_required(login_url='login')
def students(request):
    class_id = request.GET.get('class')
    sort_by = request.GET.get('sort_by', 'id')
    students = Student.objects.all().select_related('student_class')
    if class_id:
        students = students.filter(student_class_id=class_id)
    if sort_by in ['id', 'name', 'roll_number', 'created_at']:
        students = students.order_by(sort_by)
    classes = Course.objects.all()

    # Handle Add Student
    if request.method == 'POST' and request.POST.get('action') == 'add':
        name = request.POST.get('name')
        class_id_val = request.POST.get('class')
        roll_no = request.POST.get('roll_no')
        student_class = Course.objects.filter(id=class_id_val).first() if class_id_val else None
        image_file = request.FILES.get('image')
        if student_class and name and roll_no:
            student = Student.objects.create(
                name=name,
                student_class=student_class,
                roll_number=roll_no
            )
            # Set face_id to student.id and save
            student.face_id = str(student.id)
            student.save()

            # Save image if provided
            if image_file:
                folder = os.path.join(django_settings.BASE_DIR, 'face_data')
                os.makedirs(folder, exist_ok=True)
                filename = f"{student.id}.jpg"
                filepath = os.path.join(folder, filename)
                with open(filepath, 'wb+') as destination:
                    for chunk in image_file.chunks():
                        destination.write(chunk)
        return redirect(request.path + f'?class={class_id_val}&sort_by={sort_by}')

    # Handle Edit Student
    if request.method == 'POST' and request.POST.get('action') == 'edit':
        student_id = request.POST.get('student_id')
        name = request.POST.get('name')
        class_id_val = request.POST.get('class')
        roll_no = request.POST.get('roll_no')
        image_file = request.FILES.get('image')
        student = Student.objects.filter(id=student_id).first()
        if student:
            student.name = name
            student.student_class = Course.objects.filter(id=class_id_val).first() if class_id_val else None
            student.roll_number = roll_no
            student.face_id = str(student.id)
            
            # Save image if provided
            if image_file:
                folder = os.path.join(django_settings.BASE_DIR, 'face_data')
                os.makedirs(folder, exist_ok=True)
                filename = f"{student.id}.jpg"
                filepath = os.path.join(folder, filename)
                with open(filepath, 'wb+') as destination:
                    for chunk in image_file.chunks():
                        destination.write(chunk)
            student.save()
        return redirect(request.path + f'?class={class_id_val}&sort_by={sort_by}')

    # Handle Delete Student
    if request.method == 'POST' and request.POST.get('action') == 'delete':
        student_id = request.POST.get('student_id')
        student = Student.objects.filter(id=student_id).first()
        if student:
            student.delete()
        return redirect(request.path + f'?class={class_id}&sort_by={sort_by}')

    # Handle Import Students from XLSX
    if request.method == 'POST' and request.POST.get('action') == 'import_xlsx':
        xlsx_file = request.FILES.get('xlsx_file')
        import openpyxl
        from django.contrib import messages
        if xlsx_file and xlsx_file.name.endswith('.xlsx'):
            wb = openpyxl.load_workbook(xlsx_file)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            # Expecting: Name | Class | Roll No
            try:
                name_idx = header.index('Name')
                class_idx = header.index('Class')
                roll_idx = header.index('Roll No')
            except ValueError:
                messages.error(request, 'XLSX file must have columns: Name, Class, Roll No.')
                return redirect(request.path)
            added = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = row[name_idx]
                class_name = row[class_idx]
                roll_no = row[roll_idx]
                if not (name and class_name and roll_no):
                    continue
                student_class = Course.objects.filter(name=class_name).first()
                if not student_class:
                    continue
                # Prevent duplicate roll numbers in the same class
                if Student.objects.filter(student_class=student_class, roll_number=roll_no).exists():
                    continue
                student = Student.objects.create(
                    name=name,
                    student_class=student_class,
                    roll_number=roll_no
                )
                student.face_id = str(student.id)
                student.save()
                added += 1
            messages.success(request, f"Imported {added} students from XLSX file.")
        else:
            messages.error(request, 'Please upload a valid XLSX file.')
        return redirect(request.path)

    context = {
        'students': students,
        'classes': classes,
        'selected_class_id': class_id or '',
    }
    return render(request, 'contents/students.html', context)

@login_required(login_url='login')
def courses(request):
    classrooms = Course.objects.all().order_by('id')

    # Handle Add Classroom
    if request.method == 'POST' and request.POST.get('action') == 'add':
        name = request.POST.get('name')
        description = request.POST.get('description')
        if name:
            Course.objects.create(name=name, description=description)
        return redirect(request.path)

    # Handle Edit Classroom
    if request.method == 'POST' and request.POST.get('action') == 'edit':
        classroom_id = request.POST.get('classroom_id')
        name = request.POST.get('name')
        description = request.POST.get('description')
        classroom = Course.objects.filter(id=classroom_id).first()
        if classroom and name:
            classroom.name = name
            classroom.description = description
            classroom.save()
        return redirect(request.path)

    # Handle Delete Classroom
    if request.method == 'POST' and request.POST.get('action') == 'delete':
        classroom_id = request.POST.get('classroom_id')
        classroom = Course.objects.filter(id=classroom_id).first()
        if classroom:
            classroom.delete()
        return redirect(request.path)

    context = {
        'classrooms': classrooms,
    }
    return render(request, 'contents/courses.html', context)

@login_required(login_url='login')
def schedule(request):
    # Fetch all schedules, courses, and cameras
    schedules = AttendanceSchedule.objects.select_related('course', 'camera').order_by('-date', '-time')
    courses = Course.objects.all()
    cameras = Camera.objects.all()

    # Handle Add Schedule
    if request.method == 'POST' and request.POST.get('action') == 'add':
        course_id = request.POST.get('course')
        camera_id = request.POST.get('camera')
        date_val = request.POST.get('date')
        time_val = request.POST.get('time')
        if course_id and camera_id and date_val and time_val:
            AttendanceSchedule.objects.create(
                course_id=course_id,
                camera_id=camera_id,
                date=date_val,
                time=time_val
            )
        return redirect(request.path)

    # Handle Delete Schedule
    if request.method == 'POST' and request.POST.get('action') == 'delete':
        schedule_id = request.POST.get('schedule_id')
        schedule_obj = AttendanceSchedule.objects.filter(id=schedule_id).first()
        if schedule_obj:
            schedule_obj.delete()
        return redirect(request.path)

    context = {
        'schedules': schedules,
        'courses': courses,
        'cameras': cameras,
    }
    return render(request, 'contents/schedule.html', context)

@login_required(login_url='login')
def settings(request):
    return render(request, 'contents/settings.html')
